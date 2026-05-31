import openpyxl
import json
import os

wb = openpyxl.load_workbook('../Test Case SIGAP-Laravel.xlsx', data_only=True)
mapping = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        if not row: continue
        # Find something that looks like an ID (e.g. KAK-FT-001, TC-P-F01, AK-F-001, LGN-F-001, PD-F-001)
        # ID is usually the first column, but let's check first 3 columns
        for i in range(min(len(row), 3)):
            val = str(row[i]) if row[i] else ""
            import re
            if re.match(r'^[A-Z0-9-]{3,15}$', val):
                mapping[val.upper()] = sheet_name.replace('Orang Aring', 'Orang E (Dika)')
                break

# Update test-cases.json
if os.path.exists('test-cases.json'):
    with open('test-cases.json', 'r', encoding='utf-8') as f:
        cases = json.load(f)
    
    for c in cases:
        id_upper = c['id'].upper()
        tab_name = mapping.get(id_upper, 'Uncategorized')
        c['tab'] = tab_name.replace('Orang Aring', 'Orang E (Dika)')
            
    with open('test-cases.json', 'w', encoding='utf-8') as f:
        json.dump(cases, f, indent=2, ensure_ascii=False)

print(f"Updated {len(cases)} cases with tab names from Excel.")
